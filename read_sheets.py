import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

BASE = r"x:\ghostmk22\Documents\MK Cashew"

# 1. List all sheet names in Cash Book and Store
for fname in ["Cash Book 2026.xlsx", "Cashew to MK Store 2026.xlsx"]:
    wb = openpyxl.load_workbook(os.path.join(BASE, fname), data_only=False)
    print(f"\n{fname} sheets: {wb.sheetnames}")

# 2. Dump Cash Book July if it exists
wb = openpyxl.load_workbook(os.path.join(BASE, "Cash Book 2026.xlsx"), data_only=False)
july_sheets = [s for s in wb.sheetnames if 'jul' in s.lower() or 'july' in s.lower()]
print(f"\nCash Book July sheets found: {july_sheets}")
for shname in july_sheets:
    ws = wb[shname]
    print(f"\n{'='*60}")
    print(f"CASH BOOK - Sheet: {shname} (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*60}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 150)):
        row_data = [str(c.value) if c.value is not None else "" for c in row]
        if all(v == "" for v in row_data):
            continue
        while row_data and row_data[-1] == "":
            row_data.pop()
        print(" | ".join(row_data))

# If no July sheet, check last available sheet
if not july_sheets:
    print(f"\nNo July sheet found. Available: {wb.sheetnames}")
    last = wb.sheetnames[-1]
    ws = wb[last]
    print(f"\nShowing LAST sheet: {last} (rows={ws.max_row}, cols={ws.max_column})")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)):
        row_data = [str(c.value) if c.value is not None else "" for c in row]
        if all(v == "" for v in row_data):
            continue
        while row_data and row_data[-1] == "":
            row_data.pop()
        print(" | ".join(row_data))

# 3. Dump Cashew to MK Store July
wb2 = openpyxl.load_workbook(os.path.join(BASE, "Cashew to MK Store 2026.xlsx"), data_only=False)
july_sheets2 = [s for s in wb2.sheetnames if 'jul' in s.lower() or 'july' in s.lower()]
print(f"\nStore July sheets found: {july_sheets2}")
for shname in july_sheets2:
    ws = wb2[shname]
    print(f"\n{'='*60}")
    print(f"STORE - Sheet: {shname} (rows={ws.max_row}, cols={ws.max_column})")
    print(f"{'='*60}")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)):
        row_data = [str(c.value) if c.value is not None else "" for c in row]
        if all(v == "" for v in row_data):
            continue
        while row_data and row_data[-1] == "":
            row_data.pop()
        print(" | ".join(row_data))
if not july_sheets2:
    print(f"\nNo July sheet. Available: {wb2.sheetnames}")
    last = wb2.sheetnames[-1]
    ws = wb2[last]
    print(f"\nShowing LAST sheet: {last} (rows={ws.max_row})")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 100)):
        row_data = [str(c.value) if c.value is not None else "" for c in row]
        if all(v == "" for v in row_data):
            continue
        while row_data and row_data[-1] == "":
            row_data.pop()
        print(" | ".join(row_data))

# 4. Re-examine Peeling July with computed values too
print(f"\n{'='*60}")
print("PEELING JULY - Computed values (data_only=True)")
print(f"{'='*60}")
wb3 = openpyxl.load_workbook(os.path.join(BASE, "Peeling July 2026.xlsx"), data_only=True)
for shname in wb3.sheetnames:
    ws = wb3[shname]
    print(f"\n--- Sheet: {shname} ---")
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 150)):
        row_data = [str(c.value) if c.value is not None else "" for c in row]
        if all(v == "" for v in row_data):
            continue
        while row_data and row_data[-1] == "":
            row_data.pop()
        print(" | ".join(row_data))
